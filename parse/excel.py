import csv
import io
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def _is_non_empty_row(row_values: list) -> bool:
	"""Return True when at least one value in a row is meaningfully present."""
	for value in row_values:
		if value is None:
			continue
		if isinstance(value, str) and not value.strip():
			continue
		return True
	return False


def parse_excel(file_path: str) -> list[dict]:
	"""Parse any tabular file format into a list of sheet dicts."""
	suffix = Path(file_path).suffix.lower()
	try:
		if suffix in {".csv"}:
			return _parse_csv(file_path, delimiter=",")
		if suffix in {".tsv"}:
			return _parse_csv(file_path, delimiter="\t")
		if suffix == ".txt":
			return _parse_csv(file_path, delimiter="\t")
		if suffix in {".xlsx", ".xlsm"}:
			return _parse_openpyxl(file_path, read_only=False)
		if suffix == ".xls":
			return _parse_xls(file_path)
		if suffix == ".xlsb":
			return _parse_xlsb(file_path)
		if suffix == ".ods":
			return _parse_ods(file_path)
		return _parse_openpyxl(file_path, read_only=False)
	except Exception as exc:
		return [{"error": str(exc), "sheet_name": "unknown", "headers": [], "rows": []}]


def _parse_openpyxl(file_path: str, read_only: bool = False) -> list[dict]:
	"""Parse workbook formats supported by openpyxl."""
	try:
		workbook = load_workbook(filename=file_path, data_only=True, read_only=read_only)
		results: list[dict] = []
		skipped_sheet_names: list[str] = []

		for sheet_name in workbook.sheetnames:
			sheet = workbook[sheet_name]

			if not read_only:
				for merged_range in list(sheet.merged_cells.ranges):
					min_col, min_row, max_col, max_row = merged_range.bounds
					top_left_value = sheet.cell(row=min_row, column=min_col).value
					sheet.unmerge_cells(str(merged_range))
					for row_idx in range(min_row, max_row + 1):
						for col_idx in range(min_col, max_col + 1):
							sheet.cell(row=row_idx, column=col_idx).value = top_left_value

			headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())]

			data_rows: list[list] = []
			for row in sheet.iter_rows(min_row=2, values_only=True):
				row_values = list(row)
				if _is_non_empty_row(row_values):
					data_rows.append(row_values)

			if len(data_rows) == 0:
				skipped_sheet_names.append(sheet_name)
				results.append({"sheet_name": sheet_name, "skipped": True, "headers": [], "rows": [], "metadata": {"row_count": 0}})
				continue

			results.append(
				{
					"sheet_name": sheet_name,
					"headers": headers,
					"rows": data_rows,
					"metadata": {
						"row_count": len(data_rows),
						"skipped_sheets": [],
						"filename": Path(file_path).name,
					},
					"error": None,
					"skipped": False,
				}
			)

		for item in results:
			if "metadata" in item:
				item["metadata"]["skipped_sheets"] = skipped_sheet_names

		return results
	except InvalidFileException as exc:
		return [{"error": str(exc), "sheet_name": "unknown", "headers": [], "rows": []}]
	except Exception as exc:
		return [{"error": str(exc), "sheet_name": "unknown", "headers": [], "rows": []}]


def _parse_csv(file_path: str, delimiter: str = ",") -> list[dict]:
	"""Parse a CSV or TSV file into a single-sheet result."""
	try:
		encodings = ["utf-8-sig", "utf-8", "latin-1", "cp1252"]
		raw = None
		for encoding in encodings:
			try:
				with open(file_path, "r", encoding=encoding, errors="strict") as file:
					raw = file.read()
				break
			except (UnicodeDecodeError, LookupError):
				continue
		if raw is None:
			with open(file_path, "r", encoding="latin-1", errors="replace") as file:
				raw = file.read()

		reader = csv.reader(io.StringIO(raw), delimiter=delimiter)
		all_rows = [row for row in reader if any(cell.strip() for cell in row)]

		if not all_rows:
			return [
				{
					"sheet_name": Path(file_path).stem,
					"skipped": True,
					"headers": [],
					"rows": [],
					"metadata": {"row_count": 0},
				}
			]

		headers = [str(cell).strip() for cell in all_rows[0]]
		data_rows = [[str(cell).strip() for cell in row] for row in all_rows[1:]]
		data_rows = [row + [""] * max(0, len(headers) - len(row)) for row in data_rows]

		return [
			{
				"sheet_name": Path(file_path).stem,
				"headers": headers,
				"rows": data_rows,
				"metadata": {"row_count": len(data_rows), "delimiter": delimiter},
				"skipped": False,
			}
		]
	except Exception as exc:
		return [{"error": str(exc), "sheet_name": "unknown", "headers": [], "rows": []}]


def _parse_xls(file_path: str) -> list[dict]:
	"""Parse legacy .xls files using xlrd if available, openpyxl as fallback."""
	try:
		import xlrd

		workbook = xlrd.open_workbook(file_path)
		results = []
		for sheet in workbook.sheets():
			if sheet.nrows == 0:
				results.append({"sheet_name": sheet.name, "skipped": True, "headers": [], "rows": [], "metadata": {"row_count": 0}})
				continue
			headers = [str(sheet.cell_value(0, col)).strip() for col in range(sheet.ncols)]
			rows = [
				[str(sheet.cell_value(row, col)).strip() for col in range(sheet.ncols)]
				for row in range(1, sheet.nrows)
			]
			results.append(
				{
					"sheet_name": sheet.name,
					"headers": headers,
					"rows": rows,
					"metadata": {"row_count": len(rows)},
					"skipped": False,
				}
			)
		return results if results else [{"error": "no sheets found", "sheet_name": "unknown", "headers": [], "rows": []}]
	except ImportError:
		print("[WARN] xlrd not installed - attempting openpyxl for .xls file")
		return _parse_openpyxl(file_path, read_only=False)
	except Exception as exc:
		return [{"error": str(exc), "sheet_name": "unknown", "headers": [], "rows": []}]


def _parse_xlsb(file_path: str) -> list[dict]:
	"""Parse binary .xlsb files using pyxlsb if available."""
	try:
		import pyxlsb

		results = []
		with pyxlsb.open_workbook(file_path) as workbook:
			for sheet_name in workbook.sheets:
				with workbook.get_sheet(sheet_name) as sheet:
					all_rows = []
					for row in sheet.rows():
						cells = [str(item.v).strip() if item.v is not None else "" for item in row]
						if any(cells):
							all_rows.append(cells)
					if not all_rows:
						results.append({"sheet_name": sheet_name, "skipped": True, "headers": [], "rows": [], "metadata": {"row_count": 0}})
						continue
					headers = all_rows[0]
					data_rows = all_rows[1:]
					results.append(
						{
							"sheet_name": sheet_name,
							"headers": headers,
							"rows": data_rows,
							"metadata": {"row_count": len(data_rows)},
							"skipped": False,
						}
					)
		return results if results else [{"error": "no sheets", "sheet_name": "unknown", "headers": [], "rows": []}]
	except ImportError:
		print("[WARN] pyxlsb not installed - .xlsb file cannot be parsed")
		return [{"error": "pyxlsb not installed - pip install pyxlsb", "sheet_name": "unknown", "headers": [], "rows": []}]
	except Exception as exc:
		return [{"error": str(exc), "sheet_name": "unknown", "headers": [], "rows": []}]


def _parse_ods(file_path: str) -> list[dict]:
	"""Parse OpenDocument Spreadsheet .ods files using odfpy if available."""
	try:
		from odf.opendocument import load
		from odf.table import Table, TableCell, TableRow
		from odf.text import P

		document = load(file_path)
		results = []
		for sheet in document.spreadsheet.getElementsByType(Table):
			sheet_name = sheet.getAttribute("name") or "Sheet"
			all_rows = []
			for row in sheet.getElementsByType(TableRow):
				cells = []
				for cell in row.getElementsByType(TableCell):
					paragraphs = cell.getElementsByType(P)
					text = " ".join(str(paragraph) for paragraph in paragraphs).strip()
					cells.append(text)
				if any(cells):
					all_rows.append(cells)
			if not all_rows:
				results.append({"sheet_name": sheet_name, "skipped": True, "headers": [], "rows": [], "metadata": {"row_count": 0}})
				continue
			headers = all_rows[0]
			data_rows = all_rows[1:]
			results.append(
				{
					"sheet_name": sheet_name,
					"headers": headers,
					"rows": data_rows,
					"metadata": {"row_count": len(data_rows)},
					"skipped": False,
				}
			)
		return results if results else [{"error": "no sheets", "sheet_name": "unknown", "headers": [], "rows": []}]
	except ImportError:
		print("[WARN] odfpy not installed - .ods file cannot be parsed")
		return [{"error": "odfpy not installed - pip install odfpy", "sheet_name": "unknown", "headers": [], "rows": []}]
	except Exception as exc:
		return [{"error": str(exc), "sheet_name": "unknown", "headers": [], "rows": []}]
