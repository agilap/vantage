from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook


RAW_DIR = Path("data/raw")
EXCEL_DIR = RAW_DIR / "excel"
EMAIL_DIR = RAW_DIR / "emails"


def _write_revenue_table(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue"

    ws.merge_cells("A1:D1")
    ws["A1"] = "Financial Summary"

    ws.append(["Year", "Quarter", "Revenue_M", "Gross_Margin_Pct"])
    ws.append([2021, "Q1", 74.2, 42.1])
    ws.append([2021, "Q2", 78.5, 42.8])
    ws.append([2021, "Q3", 81.0, 43.0])
    ws.append([2021, "Q4", 89.4, 43.6])
    ws.append([2022, "Q1", 91.2, 44.1])
    ws.append([None, None, None, None])
    ws.append([2022, "Q2", 95.8, 44.7])
    ws.append([2022, "Q3", 99.1, 45.0])
    ws.append([2022, "Q4", 103.6, 45.4])
    ws.append([2023, "Q1", 107.3, 45.9])

    wb.save(path)


def _write_segment_breakdown(path: Path) -> None:
    wb = Workbook()

    americas = wb.active
    americas.title = "Americas"
    americas.append(["Region", "Product", "Units", "Revenue"])
    for row in [
        ["Americas", "Cloud Suite", 1200, 18_400_000],
        ["Americas", "AI Platform", 860, 14_900_000],
        ["Americas", "Security", 940, 10_700_000],
        ["Americas", "Analytics", 620, 8_100_000],
        ["Americas", "Hardware", 410, 6_600_000],
        ["Americas", "Support", 730, 5_900_000],
        ["Americas", "Services", 550, 4_800_000],
        ["Americas", "Licensing", 690, 7_200_000],
    ]:
        americas.append(row)

    emea = wb.create_sheet("EMEA")
    emea.append(["Region", "Product", "Units", "Revenue"])
    for row in [
        ["EMEA", "Cloud Suite", 980, 15_200_000],
        ["EMEA", "AI Platform", 740, 12_300_000],
        ["EMEA", "Security", 820, 9_500_000],
        ["EMEA", "Analytics", 560, 7_600_000],
        ["EMEA", "Hardware", 390, 5_900_000],
        ["EMEA", "Support", 610, 5_200_000],
        ["EMEA", "Services", 470, 4_100_000],
        ["EMEA", "Licensing", 640, 6_500_000],
    ]:
        emea.append(row)

    wb.create_sheet("Internal")
    wb.save(path)


def _write_cost_summary(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "CostSummary"
    ws.append(["Cost_Category", "Q1", "Q2", "Q3", "Q4"])

    rows = [
        ["2023 COGS", 320, 335, 342, 350],
        ["2023 R&D", 115, 118, 122, 127],
        ["2023 Sales & Marketing", 140, 145, 151, 156],
        ["2023 G&A", 60, 61, 62, 64],
        ["2023 Depreciation", 38, 39, 40, 42],
        ["2023 Other", 22, 23, 24, 24],
        ["2024 COGS", 345, 352, 360, 369],
        ["2024 R&D", 128, 133, 138, 144],
        ["2024 Sales & Marketing", 159, 163, 168, 173],
        ["2024 G&A", 65, 66, 67, 69],
        ["2024 Depreciation", 44, 45, 46, 48],
        ["2024 Other", 25, 26, 26, 27],
    ]
    for row in rows:
        ws.append(row)

    wb.save(path)


def _write_email(path: Path, content: str) -> None:
    path.write_text(content, encoding="utf-8")


def _size_kb(path: Path) -> float:
    return path.stat().st_size / 1024.0


def main() -> int:
    EXCEL_DIR.mkdir(parents=True, exist_ok=True)
    EMAIL_DIR.mkdir(parents=True, exist_ok=True)

    created_files: list[Path] = []

    revenue_table = EXCEL_DIR / "revenue_table.xlsx"
    _write_revenue_table(revenue_table)
    created_files.append(revenue_table)

    segment_breakdown = EXCEL_DIR / "segment_breakdown.xlsx"
    _write_segment_breakdown(segment_breakdown)
    created_files.append(segment_breakdown)

    cost_summary = EXCEL_DIR / "cost_summary.xlsx"
    _write_cost_summary(cost_summary)
    created_files.append(cost_summary)

    q3_thread = EMAIL_DIR / "q3_projections.eml"
    _write_email(
        q3_thread,
        "Subject: Q3 Projection Update\n"
        "From: maya.chen@vantage.example\n"
        "Date: Tue, 09 Jul 2024 09:00:00 -0500\n"
        "\n"
        "We are tracking Q3 revenue at $412M, slightly above the prior forecast of $405M. "
        "Enterprise renewals are ahead of plan and upsell pipeline conversion improved in June.\n"
        "\n"
        "From: daniel.ruiz@vantage.example\n"
        "Date: Tue, 09 Jul 2024 11:20:00 -0500\n"
        "\n"
        "Thanks, I agree with the upward revision and suggest we lock guidance at $410M to preserve buffer.\n"
        "\n"
        "Quoted message:\n"
        "> We are tracking Q3 revenue at $412M, slightly above the prior forecast of $405M.\n"
        "> Enterprise renewals are ahead of plan and upsell pipeline conversion improved in June.\n",
    )
    created_files.append(q3_thread)

    supply_chain = EMAIL_DIR / "supply_chain_update.eml"
    _write_email(
        supply_chain,
        "Subject: Supply Chain Update - Component Availability\n"
        "From: rachel.nguyen@vantage.example\n"
        "Date: Mon, 12 Aug 2024 08:15:00 -0500\n"
        "\n"
        "Our latest supplier check-in shows continued shortages in high-density memory modules from Micron "
        "and power management ICs from Infineon. Current inbound allocations remain below committed volumes for "
        "the next two production cycles.\n"
        "\n"
        "Lead times for the affected assemblies have stretched from six weeks to nine weeks, with the largest "
        "impact on premium configuration SKUs. Teams should prioritize customer orders that include contractual "
        "delivery penalties and defer low-margin bundles until material flow normalizes.\n"
        "\n"
        "Procurement is negotiating secondary coverage with Texas Instruments distributors and evaluating alternate "
        "BOM paths for two controller boards. We expect mitigation actions to recover approximately 35% of the "
        "current gap by month-end if substitute qualification completes on schedule.\n",
    )
    created_files.append(supply_chain)

    board_summary = EMAIL_DIR / "board_summary.eml"
    _write_email(
        board_summary,
        "Subject: Board Summary\n"
        "From: ceo.office@vantage.example\n"
        "Date: Fri, 30 Aug 2024 17:45:00 -0500\n"
        "\n"
        "Board packet approved.\n",
    )
    created_files.append(board_summary)

    rd_budget = EMAIL_DIR / "rd_budget.eml"
    _write_email(
        rd_budget,
        "Subject: FY2024 R&D Budget Approval\n"
        "From: cfo@vantage.example\n"
        "Date: Wed, 04 Sep 2024 10:05:00 -0500\n"
        "\n"
        "Approved: total R&D budget for fiscal year 2024 is $186,000,000, with $92,000,000 allocated to platform "
        "engineering and $58,000,000 allocated to applied AI product work. The remaining $36,000,000 will support "
        "compliance, reliability, and security initiatives across shared services.\n"
        "\n"
        "Please align quarterly spend plans to this envelope and submit variance justifications for any team "
        "projecting more than a 3% deviation. Finance will track burn against a quarterly baseline of $46,500,000 "
        "and publish dashboard updates after each month close.\n",
    )
    created_files.append(rd_budget)

    print("Created files")
    print("| Path | Size (KB) |")
    print("|---|---:|")
    for path in created_files:
        print(f"| {path.as_posix()} | {_size_kb(path):.1f} |")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
